from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font, NamedStyle
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from django.http import HttpResponse
from django.utils.http import urlquote
from wsgiref.util import FileWrapper

import os
import shutil
import unicodedata


class ExcelExport(object):
    def __init__(self, font_name=None):
        self.font_name = font_name or 'Arial'

        self.CENTER_ALIGNMENT = Alignment(vertical='top', horizontal='center')
        self.LEFT_ALIGNMENT = Alignment(vertical='top', horizontal='left')
        self.RIGHT_ALIGNMENT = Alignment(vertical='top', horizontal='right')

        self.INTEGER_FORMAT = '_(* #,##0_);_(* (#,##0);_(* "-"??_);_(@_)'
        self.WEIGHT_FORMAT = '_(* #,##0.000_);_(* (#,##0.000);_(* "-"??_);_(@_)'
        self.CBM_FORMAT = '_(* #,##0.0000_);_(* (#,##0.0000);_(* "-"??_);_(@_)'
        self.DATE_FORMAT = 'yyyy-mm-dd'
        self.TEXT_FORMAT = '@'

        self.DOLLAR_FORMAT = u'_($RM* #,##0.00_);[Red]_($RM* (#,##0.00);_($RM* -_0_0_);_(@'
        self.CURRENCY_FORMAT = u'* #,##0.00;[Red]* (#,##0.00);* -;@'

        self.ITEM_FONT = Font(name=self.font_name, size=10)
        self.ITEM_BOLD_FONT = Font(name=self.font_name, size=10, bold=True)

        self.BORDER_SUMMARY = Border(bottom=Side(border_style='double'))
        self.BORDER_CELL = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        self.BORDER_TOP = Border(top=Side(border_style=BORDER_THIN))
        self.BORDER_LEFT = Border(left=Side(border_style=BORDER_THIN))
        self.BORDER_RIGHT = Border(right=Side(border_style=BORDER_THIN))
        self.BORDER_BOTTOM = Border(bottom=Side(border_style=BORDER_THIN))

        self.DATE_STYLE_BORDER = NamedStyle(name='date_style_border')
        self.DATE_STYLE_BORDER.font = self.ITEM_FONT
        self.DATE_STYLE_BORDER.number_format = self.DATE_FORMAT
        self.DATE_STYLE_BORDER.alignment = self.CENTER_ALIGNMENT
        self.DATE_STYLE_BORDER.border = self.BORDER_CELL

        self.DATE_STYLE = NamedStyle(name='date_style')
        self.DATE_STYLE.font = self.ITEM_FONT
        self.DATE_STYLE.number_format = self.DATE_FORMAT
        self.DATE_STYLE.alignment = self.CENTER_ALIGNMENT

        self.TITLE_HEADER = NamedStyle(name='TITLE_HEADER')
        self.TITLE_HEADER.font = Font(name='Arial', size=20, bold=True, color='4472C4')
        self.TITLE_HEADER.alignment = Alignment(vertical='center', horizontal='center')

        self.PRICE_HEADER_ONE = NamedStyle(name='PRICE_HEADER')
        self.PRICE_HEADER_ONE.font = Font(name='Arial', size=16, bold=True)
        self.PRICE_HEADER_ONE.alignment = Alignment(vertical='center', horizontal='center')
        self.PRICE_HEADER_ONE.number_format = self.CURRENCY_FORMAT

        self.HEADER_CAPTION = NamedStyle(name='header_caption')
        self.HEADER_CAPTION.font = Font(name='SimHei', size=10, bold=True)
        self.HEADER_CAPTION.number_format = '@'
        self.HEADER_CAPTION.alignment = Alignment(vertical='top', horizontal='right')
        self.HEADER_CAPTION.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        self.HEADER_CAPTION.fill = PatternFill(start_color='BFBFBF', fill_type='solid')

        self.HEADER_DATA = NamedStyle(name='header_data')
        self.HEADER_DATA.font = Font(name=self.font_name, size=10, bold=False)
        self.HEADER_DATA.number_format = '@'
        self.HEADER_DATA.alignment = Alignment(vertical='top', horizontal='left')
        self.HEADER_DATA.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        self.TEXT_CENTER_STYLE_BORDER = NamedStyle(name='text_center_style_border')
        self.TEXT_CENTER_STYLE_BORDER.font = self.ITEM_FONT
        self.TEXT_CENTER_STYLE_BORDER.number_format = self.TEXT_FORMAT
        self.TEXT_CENTER_STYLE_BORDER.alignment = self.CENTER_ALIGNMENT
        self.TEXT_CENTER_STYLE_BORDER.border = self.BORDER_CELL

        self.TEXT_LEFT_STYLE_BORDER = NamedStyle(name='text_left_style_border')
        self.TEXT_LEFT_STYLE_BORDER.font = self.ITEM_FONT
        self.TEXT_LEFT_STYLE_BORDER.number_format = self.TEXT_FORMAT
        self.TEXT_LEFT_STYLE_BORDER.alignment = self.LEFT_ALIGNMENT
        self.TEXT_LEFT_STYLE_BORDER.border = self.BORDER_CELL

        self.TEXT_CENTER_STYLE = NamedStyle(name='text_center_style')
        self.TEXT_CENTER_STYLE.font = self.ITEM_FONT
        self.TEXT_CENTER_STYLE.number_format = self.TEXT_FORMAT
        self.TEXT_CENTER_STYLE.alignment = self.CENTER_ALIGNMENT

        self.TEXT_LEFT_STYLE = NamedStyle(name='text_left_style')
        self.TEXT_LEFT_STYLE.font = self.ITEM_FONT
        self.TEXT_LEFT_STYLE.number_format = self.TEXT_FORMAT
        self.TEXT_LEFT_STYLE.alignment = self.LEFT_ALIGNMENT

        self.QUANTITY_STYLE = NamedStyle(name='quantity_style')
        self.QUANTITY_STYLE.font = self.ITEM_FONT
        self.QUANTITY_STYLE.number_format = self.INTEGER_FORMAT
        self.QUANTITY_STYLE.alignment = self.RIGHT_ALIGNMENT
        self.QUANTITY_STYLE.border = self.BORDER_CELL

        self.DECIMAL_3_STYLE = NamedStyle(name='decimal_3_style')
        self.DECIMAL_3_STYLE.font = self.ITEM_FONT
        self.DECIMAL_3_STYLE.number_format = self.WEIGHT_FORMAT
        self.DECIMAL_3_STYLE.alignment = self.RIGHT_ALIGNMENT
        self.DECIMAL_3_STYLE.border = self.BORDER_CELL

        self.DECIMAL_4_STYLE = NamedStyle(name='decimal_4_style')
        self.DECIMAL_4_STYLE.font = self.ITEM_FONT
        self.DECIMAL_4_STYLE.number_format = self.CBM_FORMAT
        self.DECIMAL_4_STYLE.alignment = self.RIGHT_ALIGNMENT
        self.DECIMAL_4_STYLE.border = self.BORDER_CELL

        self.COLUMN_TITLE_CN = NamedStyle(name='column_title_cn')
        self.COLUMN_TITLE_CN.font = Font(name='SimHei', size=10, bold=False)
        self.COLUMN_TITLE_CN.number_format = self.TEXT_FORMAT
        self.COLUMN_TITLE_CN.alignment = self.CENTER_ALIGNMENT
        self.COLUMN_TITLE_CN.border = self.BORDER_CELL
        self.COLUMN_TITLE_CN.fill = PatternFill(start_color='BFBFBF', fill_type='solid')

        self.COLUMN_TITLE_EN = NamedStyle(name='column_title_en')
        self.COLUMN_TITLE_EN.font = Font(name=self.font_name, size=10, bold=True)
        self.COLUMN_TITLE_EN.number_format = self.TEXT_FORMAT
        self.COLUMN_TITLE_EN.alignment = self.CENTER_ALIGNMENT
        self.COLUMN_TITLE_EN.border = self.BORDER_CELL
        self.COLUMN_TITLE_EN.fill = PatternFill(start_color='BFBFBF', fill_type='solid')

    def draw_border(self, worksheet, row, col, width=None, height=None):
        side = Side(style='thin')

        if not width:
            width = 1

        if not height:
            height = 1

        if width == 1 and height == 1:
            worksheet.cell(row=row, column=col).border = self.BORDER_CELL
        elif width == 1 and height == 2:
            worksheet.cell(row=row, column=col).border = Border(left=side, right=side, top=side)
            worksheet.cell(row=row + 1, column=col).border = Border(left=side, right=side, bottom=side)
        elif width == 1 and height >= 3:
            for i in range(0, height):
                if i == 0:
                    worksheet.cell(row=row, column=col).border = Border(left=side, right=side, top=side)
                elif i == height - 1:
                    worksheet.cell(row=row + i, column=col).border = Border(left=side, right=side, bottom=side)
                else:
                    worksheet.cell(row=row + i, column=col).border = Border(left=side, right=side)

    def remove_path(self, path):
        if os.path.exists(path):
            if os.path.isdir(path):
                shutil.rmtree(path)
            else:
                os.remove(path)

    def clone(self, from_template, to_export):
        os.makedirs(os.path.dirname(to_export), exist_ok=True)

        self.remove_path(to_export)
        shutil.copy(from_template, to_export)
        workbook = load_workbook(to_export)
        return workbook

    def rfc5987_content_disposition(self, filename):
        ascii_name = unicodedata.normalize('NFKD', filename).encode('ascii', 'ignore').decode()
        header = 'attachment; filename="{}"'.format(ascii_name)
        if ascii_name != filename:
            quoted_name = urlquote(filename)
            header += '; filename*=UTF-8\'\'{}'.format(quoted_name)

        return header

    def to_response(self, export_file_path, download_filename):
        file = open(export_file_path, 'rb')

        response = HttpResponse(FileWrapper(file), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = self.rfc5987_content_disposition(download_filename)

        return response
